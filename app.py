from flask import Flask, render_template, send_file, Response, request
from openpyxl import Workbook
import requests
import cachetools.func
import yaml
import dicttoxml

from xml.dom.minidom import parseString
import datetime
import json
import csv
import os
import io
import string

with open('configs.json') as file:
    configs = json.load(file)

ttl = configs['ttl'] if 'ttl' in configs.keys() else 0

app = Flask(__name__)


def get_archive_readings(request):
    try:
        start_date = request.args.get("start_date")
        end_date = request.args.get("end_date")

        start_date_dt = datetime.datetime.fromisoformat(start_date)
        end_date_dt = datetime.datetime.fromisoformat(end_date)

        start_date_dt = datetime.datetime(year=start_date_dt.year,
                                            month=start_date_dt.month,
                                            day=start_date_dt.day,
                                            hour=0,
                                            minute=0)
        end_date_dt = datetime.datetime(year=end_date_dt.year,
                                        month=end_date_dt.month,
                                        day=end_date_dt.day,
                                        hour=23,
                                        minute=59)
        end_date_dt += datetime.timedelta(days=1)
    except Exception:
        start_date = None
        end_date = None

        start_date_dt = None
        end_date_dt = None

    all = request.args.get('all') == 'true'

    status = 200

    if (start_date is None or end_date is None) and not all:
        status = 400
        rd = {
            'status': 'error',
            'error': 'start_date_and_end_date_required'
        }

    if status == 200:
        try:
            if not all:
                r = requests.get(f'{configs["remote_url"]}/api/archive_readings',
                                params={
                                    'startDate': start_date_dt.isoformat(),
                                    'endDate': end_date_dt.isoformat()
                                })
            else:
                r = requests.get(f'{configs["remote_url"]}/api/archive_readings')
        except Exception:
            rd = {
                'status': 'error',
                'error': 'internal_error'
            }
            status = 500

    if status == 200:
        if r.status_code == 200:
            rj = r.json()
        else:
            rd = {
                'status': 'error',
                'error': 'internal_error'
            }
            status = 500

    if status == 200:
        readings = {}

        for r in rj:
            dt = datetime.datetime.fromisoformat(r['read_time'])
            dt = datetime.datetime.utcfromtimestamp(dt.timestamp())
            dt = datetime.datetime(year=dt.year,
                                   month=dt.month,
                                   day=dt.day,
                                   hour=dt.hour)
            dt_iso = dt.isoformat()

            if dt_iso not in readings.keys():
                readings[dt_iso] = {
                    'temperature': [],
                    'humidity': [],
                    'pressure': [],
                    'pm1.0': [],
                    'pm2.5': [],
                    'pm10': []
                }

            if r['ds18b20_temperature'] is not None:
                readings[dt_iso]['temperature'].append(r['ds18b20_temperature'])
            if r['bme280_humidity'] is not None:
                readings[dt_iso]['humidity'].append(r['bme280_humidity'])
            if r['bme280_pressure'] is not None:
                readings[dt_iso]['pressure'].append(r['bme280_pressure'])
            if r['pms5003_pm_1_0'] is not None:
                readings[dt_iso]['pm1.0'].append(r['pms5003_pm_1_0'])
            if r['pms5003_pm_2_5'] is not None:
                readings[dt_iso]['pm2.5'].append(r['pms5003_pm_2_5'])
            if r['pms5003_pm_10'] is not None:
                readings[dt_iso]['pm10'].append(r['pms5003_pm_10'])

        for date in readings.keys():
            for key in ['temperature', 'humidity', 'pressure', 'pm1.0', 'pm2.5', 'pm10']:
                if len(readings[date][key]) != 0:
                    readings[date][key] = sum(readings[date][key]) / len(readings[date][key])
                else:
                    readings[date][key] = None

        for date in readings.keys():
            if readings[date]['humidity'] is not None and readings[date]['temperature'] is not None:
                readings[date]['dewpoint'] = ((readings[date]['humidity'] / 100) ** (1 / 8)) * (112 + (0.9 * readings[date]['temperature'])) + (0.1 * readings[date]['temperature']) - 112
            else:
                readings[date]['dewpoint'] = None
            if readings[date]['pm10'] is not None and readings[date]['pm2.5'] is not None:
                readings[date]['aqi'] = ((readings[date]['pm10'] / 1.8) + (readings[date]['pm2.5'] / 1.1)) / 2
            else:
                readings[date]['aqi'] = None

        for date in readings.keys():
            readings[date]['temperature'] = round(readings[date]['temperature'], 1) if readings[date]['temperature'] is not None else None
            readings[date]['humidity'] = round(readings[date]['humidity']) if readings[date]['humidity'] is not None else None
            readings[date]['pressure'] = round(readings[date]['pressure']) if readings[date]['pressure'] is not None else None
            readings[date]['dewpoint'] = round(readings[date]['dewpoint'], 1) if readings[date]['dewpoint'] is not None else None
            readings[date]['pm1.0'] = round(readings[date]['pm1.0']) if readings[date]['pm1.0'] is not None else None
            readings[date]['pm2.5'] = round(readings[date]['pm2.5']) if readings[date]['pm2.5'] is not None else None
            readings[date]['pm10'] = round(readings[date]['pm10']) if readings[date]['pm10'] is not None else None
            readings[date]['aqi'] = round(readings[date]['aqi']) if readings[date]['aqi'] is not None else None

        readings2 = []

        for key, value in readings.items():
            if all or datetime.datetime.fromisoformat(start_date).date() <= \
                datetime.datetime.fromisoformat(key).date() <= \
                datetime.datetime.fromisoformat(end_date).date():
                readings2.append(value)
                readings2[-1]['date'] = key

    if status == 200:
        rd = {
            'status': 'ok',
            'readings': readings2
        }

    res = Response(
        json.dumps(rd, indent=4),
        status=status,
        mimetype='application/json'
    )

    res.headers['Access-Control-Allow-Origin'] = '*'
    return res


@app.route('/')
def index():
    return render_template('index.html', configs=configs)


@app.route('/widget')
def widget():
    return render_template('widget.html')


@app.route('/manifest.json')
def manifest():
    return send_file('static/manifest.json')


@app.route('/manifest_en.json')
def manifest_en():
    return send_file('static/manifest_en.json')


@app.route('/sw.js')
def sw():
    return send_file('static/js/sw.js')


@app.route('/favicon.ico')
def favicon():
    return send_file('static/img/icon.ico')


@app.route('/api/current_readings')
@cachetools.func.ttl_cache(maxsize=128, ttl=ttl)
def current_readings():
    status = 200

    try:
        r = requests.get(f'{configs["remote_url"]}/api/current_reading')
    except Exception:
        rd = {
            'status': 'error'
        }
        status = 500

    if status == 200:
        if r.status_code == 200:
            rj = r.json()
        else:
            rd = {
                'status': 'error'
            }
            status = 500

    if status == 200:
        if rj['status'] == 'error':
            rd = {
                'status': 'error'
            }
            status = 500
        elif rj['status'] == 'error':
            rd = {
                'status': 'error'
            }
            status = 500

    if status == 200:
        rd = {
            'status': 'ok'
        }

        dt = datetime.datetime.fromisoformat(rj['date'])
        dt_utc = datetime.datetime.utcfromtimestamp(dt.timestamp())
        rd['date'] = dt_utc.isoformat()

        try:
            if rj['ds18b20']['temperature'] is not None:
                rd['temperature'] = round(rj['ds18b20']['temperature'], 1)
            else:
                rd['temperature'] = None
        except Exception:
            rd['temperature'] = None

        try:
            if rj['bme280']['humidity'] is not None:
                rd['humidity'] = round(rj['bme280']['humidity'])
            else:
                rd['humidity'] = None
        except Exception:
            rd['humidity'] = None

        try:
            if rj['bme280']['pressure'] is not None:
                rd['pressure'] = round(rj['bme280']['pressure'])
            else:
                rd['pressure'] = None
        except Exception:
            rd['pressure'] = None

        try:
            if rj['ds18b20']['temperature'] is not None and rj['bme280']['humidity'] is not None:
                rd['dewpoint'] = round(((rj['bme280']['humidity'] / 100) ** (1 / 8)) * (112 + (0.9 * rj['ds18b20']['temperature'])) + (0.1 * rj['ds18b20']['temperature']) - 112, 1)
            else:
                rd['dewpoint'] = None
        except Exception:
            rd['dewpoint'] = None

        try:
            if rj['pms5003']['pm1.0'] is not None:
                rd['pm1.0'] = rj['pms5003']['pm1.0']
            else:
                rd['pm1.0'] = None
        except Exception:
            rd['pm1.0'] = None

        try:
            if rj['pms5003']['pm2.5'] is not None:
                rd['pm2.5'] = rj['pms5003']['pm2.5']
            else:
                rd['pm2.5'] = None
        except Exception:
            rd['pm2.5'] = None

        try:
            if rj['pms5003']['pm10'] is not None:
                rd['pm10'] = rj['pms5003']['pm10']
            else:
                rd['pm10'] = None
        except Exception:
            rd['pm10'] = None

        try:
            if rj['pms5003']['pm10'] is not None and rj['pms5003']['pm2.5'] is not None:
                rd['aqi'] = round(((rj['pms5003']['pm10'] / 1.8) + (rj['pms5003']['pm2.5'] / 1.1)) / 2)
            else:
                rd['aqi'] = None
        except Exception:
            rd['aqi'] = None

    res = Response(
        json.dumps(rd, indent=4),
        status=status,
        mimetype='application/json'
    )

    res.headers['Access-Control-Allow-Origin'] = '*'
    return res


@app.route('/api/archive_readings')
def archive_readings():
    return get_archive_readings(request)


@app.route('/api/archive_readings/download/json')
def archive_readings_json():
    r = get_archive_readings(request)
    rs = r.status_code
    rj = r.json

    if rs == 200:
        indent = 4 if request.args.get('format') == 'true' else None
        readings = rj['readings']
        file = io.StringIO()
        json.dump(readings, file, indent=indent)

        file_bytes = io.BytesIO()
        file_bytes.write(file.getvalue().encode())
        file_bytes.seek(0)
        file.close()

        return send_file(
            file_bytes,
            as_attachment=True,
            download_name='readings.json',
            mimetype='application/json'
        )

    return "Error", 400


@app.route('/api/archive_readings/download/csv')
def archive_readings_csv():
    r = get_archive_readings(request)
    rs = r.status_code
    rj = r.json

    field_names = ["date", "temperature", "humidity", "pressure", "dewpoint",
                   "pm1.0", "pm2.5", "pm10", "aqi"]

    if rs == 200:
        readings = rj['readings']
        file = io.StringIO()

        writer = csv.DictWriter(file, fieldnames=field_names)
        writer.writeheader()
        writer.writerows(readings)

        file_bytes = io.BytesIO()
        file_bytes.write(file.getvalue().encode())
        file_bytes.seek(0)
        file.close()

        return send_file(
            file_bytes,
            as_attachment=True,
            download_name='readings.csv',
            mimetype='text/csv'
        )

    return "Error", 400


@app.route('/api/archive_readings/download/sql')
def archive_readings_sql():
    r = get_archive_readings(request)
    rs = r.status_code
    rj = r.json

    if rs == 200:
        readings = rj['readings']
        readings_count = len(readings)
        file = io.StringIO()

        file.write('''SET SQL_MODE = "NO_AUTO_VALUE_ON_ZERO";
START TRANSACTION;
SET time_zone = "+00:00";


/*!40101 SET @OLD_CHARACTER_SET_CLIENT=@@CHARACTER_SET_CLIENT */;
/*!40101 SET @OLD_CHARACTER_SET_RESULTS=@@CHARACTER_SET_RESULTS */;
/*!40101 SET @OLD_COLLATION_CONNECTION=@@COLLATION_CONNECTION */;
/*!40101 SET NAMES utf8mb4 */;


CREATE TABLE `readings` (
  `id` int,
  `date` datetime,
  `temperature` float,
  `humidity` int,
  `pressure` int,
  `dewpoint` float,
  `pm1_0` int,
  `pm2_5` int,
  `pm10` int,
  `aqi` int
) ENGINE=InnoDB DEFAULT CHARSET=utf8;


ALTER TABLE `readings`
  ADD PRIMARY KEY (`id`);


ALTER TABLE `readings`
  MODIFY `id` int NOT NULL AUTO_INCREMENT;
COMMIT;


INSERT INTO `readings` (`date`, `temperature`, `humidity`, `pressure`, `dewpoint`, `pm1_0`, `pm2_5`, `pm10`, `aqi`) VALUES
''')

        for i, reading in enumerate(readings):
            date = datetime.datetime.fromisoformat(reading['date']).strftime("%Y-%m-%d %H:%M:%S")
            temperature = str(reading['temperature']) if reading['temperature'] is not None else 'NULL'
            humidity = str(reading['humidity']) if reading['humidity'] is not None else 'NULL'
            pressure = str(reading['pressure']) if reading['pressure'] is not None else 'NULL'
            dewpoint = str(reading['dewpoint']) if reading['dewpoint'] is not None else 'NULL'
            pm1_0 = str(reading['pm1.0']) if reading['pm1.0'] is not None else 'NULL'
            pm2_5 = str(reading['pm2.5']) if reading['pm2.5'] is not None else 'NULL'
            pm10 = str(reading['pm10']) if reading['pm10'] is not None else 'NULL'
            aqi = str(reading['aqi']) if reading['aqi'] is not None else 'NULL'

            file.write(f"('{date}', {temperature}, {humidity}, {pressure}, {dewpoint}, {pm1_0}, {pm2_5}, {pm10}, {aqi})")

            if i + 1 == readings_count:
                file.write(';')
            else:
                file.write(',\n')

        file.write('''


/*!40101 SET CHARACTER_SET_CLIENT=@OLD_CHARACTER_SET_CLIENT */;
/*!40101 SET CHARACTER_SET_RESULTS=@OLD_CHARACTER_SET_RESULTS */;
/*!40101 SET COLLATION_CONNECTION=@OLD_COLLATION_CONNECTION */;
''')

        file_bytes = io.BytesIO()
        file_bytes.write(file.getvalue().encode())
        file_bytes.seek(0)
        file.close()

        return send_file(
            file_bytes,
            as_attachment=True,
            download_name='readings.sql',
            mimetype='text/plain'
        )

    return "Error", 400


@app.route('/api/archive_readings/download/yaml')
def archive_readings_yaml():
    r = get_archive_readings(request)
    rs = r.status_code
    rj = r.json

    if rs == 200:
        readings = rj['readings']
        file = io.StringIO()

        yaml.dump(readings, file)

        file_bytes = io.BytesIO()
        file_bytes.write(file.getvalue().encode())
        file_bytes.seek(0)
        file.close()

        return send_file(
            file_bytes,
            as_attachment=True,
            download_name='readings.yaml',
            mimetype='text/yaml'
        )

    return "Error", 400


@app.route('/api/archive_readings/download/xml')
def archive_readings_xml():
    r = get_archive_readings(request)
    rs = r.status_code
    rj = r.json

    if rs == 200:
        readings = rj['readings']
        file = io.StringIO()

        xml = dicttoxml.dicttoxml(readings,
                                  custom_root='readings',
                                  attr_type=False,
                                  item_func=lambda x: 'reading')

        if request.args.get('format') == 'true':
            xml = parseString(xml).toprettyxml()
            xml = xml.replace('\t', '    ')
        else:
            xml = xml.decode()

        file.write(xml)

        file_bytes = io.BytesIO()
        file_bytes.write(file.getvalue().encode())
        file_bytes.seek(0)
        file.close()

        return send_file(
            file_bytes,
            as_attachment=True,
            download_name='readings.xml',
            mimetype='text/xml'
        )

    return "Error", 400


@app.route('/api/archive_readings/download/excel')
def archive_readings_excel():
    r = get_archive_readings(request)
    rs = r.status_code
    rj = r.json

    names = {"date": "Date",
             "temperature": "Temperature",
             "humidity": "Humidity",
             "pressure": "Pressure",
             "dewpoint": "Dewpoint",
             "pm1.0": "PM 1.0",
             "pm2.5": "PM 2.5",
             "pm10": "PM 10",
             "aqi": "AQI"}

    if rs == 200:
        readings = rj['readings']

        wb = Workbook()
        wb.properties.creator = "Swarzędz Meteo"
        ws = wb.active

        ws.title = "Readings"

        for i, name in enumerate(names.values()):
            ws[string.ascii_uppercase[i] + '1'] = name

        for i, reading in enumerate(readings):
            for j, name in enumerate(names.keys()):
                ws[string.ascii_uppercase[j] + str(i + 2)] = reading[name]

        file_bytes = io.BytesIO()

        wb.save(file_bytes)

        file_bytes.seek(0)
        file.close()

        return send_file(
            file_bytes,
            as_attachment=True,
            download_name='readings.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    return "Error", 400


@app.route('/api/announcements')
def announcements():
    if os.path.isfile('announcements.json'):
        try:
            with open('announcements.json', encoding='utf-8') as file:
                announcements = json.load(file)

            res = Response(
                json.dumps({
                    'status': 'ok',
                    'announcements': announcements
                }),
                status=200,
                mimetype='application/json'
            )
        except Exception:
            res = Response(
                json.dumps({
                    'status': 'error'
                }),
                status=500,
                mimetype='application/json'
            )
    else:
        res = Response(
            json.dumps({
                'status': 'ok',
                'announcements': []
            }),
            status=200,
            mimetype='application/json'
        )

    res.headers['Access-Control-Allow-Origin'] = '*'
    return res


@app.route('/api/status')
@cachetools.func.ttl_cache(maxsize=128, ttl=ttl)
def status():
    try:
        r = requests.get(f'{configs["remote_url"]}/api/current_reading')
        station_up = True
    except Exception:
        station_up = False
    
    if station_up:
        if r.status_code == 200:
            rj = r.json()
        else:
            station_up = False
    
    if station_up:
        ds18b20 = rj['ds18b20']['temperature'] is not None
        bme280 = rj['bme280']['humidity'] is not None and \
            rj['bme280']['pressure'] is not None
        pms5003 = rj['pms5003']['pm1.0'] is not None and \
            rj['pms5003']['pm2.5'] is not None and \
            rj['pms5003']['pm10'] is not None
        everything = station_up and ds18b20 and bme280 and pms5003
    else:
        ds18b20 = False
        bme280 = False
        pms5003 = False
        everything = False

    res = Response(
        json.dumps({
            'status': 'ok',
            'station': station_up,
            'ds18b20': ds18b20,
            'bme280': bme280,
            'pms5003': pms5003,
            'everything': everything
        }),
        status=200,
        mimetype='application/json'
    )

    res.headers['Access-Control-Allow-Origin'] = '*'
    return res


if __name__ == '__main__':
    host = configs['host'] if 'host' in configs.keys() else None
    port = configs['port'] if 'port' in configs.keys() else None
    debug = configs['debug'] if 'debug' in configs.keys() else None

    app.run(host=host, port=port, debug=debug)
